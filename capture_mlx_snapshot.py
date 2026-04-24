#!/usr/bin/env python3
"""
Capture Arduino-style MLX90394 snapshot blocks from either a serial port or a TCP socket,
and save them to CSV. Also provides helpers used by scan_mlx_z_stack.py.

Supported sources:
- Serial COM port, e.g. COM7
- TCP source, e.g. tcp://127.0.0.1:8765

The TCP mode is intended for dummy/test data generators that emit the same text snapshot
format as the Arduino sketch.

Requirements on Windows:
    py -m pip install pyserial openpyxl
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import pathlib
import re
import socket
import sys
import time
from typing import Dict, Iterable, List, Protocol, Tuple
from urllib.parse import urlparse

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl\n"
        "Install with: py -m pip install openpyxl"
    ) from exc

try:
    import serial
    from serial.tools import list_ports
except ImportError:
    serial = None
    list_ports = None


SCRIPT_DIR = pathlib.Path(__file__).resolve().parent

TCA1_CHANNELS = range(0, 7)
TCA2_CHANNELS = range(0, 8)
MLX_ADDRS = (0x60, 0x61)
START_END_MARKER_PREFIX = "###"
STATUS_ONLY = {"NOTCA2", "ABSENT", "CFGERR", "NOSAMP", "RDERR"}
WIDE_BASE_FIELDS = [
    "z_index",
    "z_steps",
    "z_um",
    "z_mm",
    "snapshot_arduino_ms",
    "capture_pc_local_time",
    "sensors_present",
    "sensors_measured",
]


class SerialCommandError(RuntimeError):
    pass


class SnapshotSource(Protocol):
    source_name: str

    def readline(self) -> bytes:
        ...

    def write(self, payload: bytes) -> int:
        ...

    def flush(self) -> None:
        ...

    def reset_input_buffer(self) -> None:
        ...

    def reset_output_buffer(self) -> None:
        ...

    def close(self) -> None:
        ...


class SerialSnapshotSource:
    def __init__(self, port: str, baud: int, settle_s: float) -> None:
        self._ser = serial.Serial(port=port, baudrate=baud, timeout=0.25, write_timeout=2)
        self.source_name = port
        time.sleep(settle_s)
        self._ser.reset_input_buffer()
        self._ser.reset_output_buffer()

    def readline(self) -> bytes:
        return self._ser.readline()

    def write(self, payload: bytes) -> int:
        return self._ser.write(payload)

    def flush(self) -> None:
        self._ser.flush()

    def reset_input_buffer(self) -> None:
        self._ser.reset_input_buffer()

    def reset_output_buffer(self) -> None:
        self._ser.reset_output_buffer()

    def close(self) -> None:
        self._ser.close()


class TcpSnapshotSource:
    def __init__(self, host: str, port: int, settle_s: float) -> None:
        self._sock = socket.create_connection((host, port), timeout=5.0)
        self._sock.settimeout(0.25)
        self._buffer = bytearray()
        self.source_name = f"tcp://{host}:{port}"
        if settle_s > 0:
            time.sleep(settle_s)
        self.reset_input_buffer()

    def readline(self) -> bytes:
        deadline = time.monotonic() + 0.25
        while time.monotonic() < deadline:
            newline_pos = self._buffer.find(b"\n")
            if newline_pos >= 0:
                line = bytes(self._buffer[: newline_pos + 1])
                del self._buffer[: newline_pos + 1]
                return line
            try:
                chunk = self._sock.recv(4096)
            except socket.timeout:
                continue
            if not chunk:
                return b""
            self._buffer.extend(chunk)
        return b""

    def write(self, payload: bytes) -> int:
        self._sock.sendall(payload)
        return len(payload)

    def flush(self) -> None:
        return None

    def reset_input_buffer(self) -> None:
        self._buffer.clear()
        self._sock.setblocking(False)
        try:
            while True:
                chunk = self._sock.recv(4096)
                if not chunk:
                    break
        except (BlockingIOError, socket.timeout):
            pass
        finally:
            self._sock.setblocking(True)
            self._sock.settimeout(0.25)

    def reset_output_buffer(self) -> None:
        return None

    def close(self) -> None:
        self._sock.close()


def normalize_source(source: str | None, port: str | None) -> str:
    if source:
        return source
    if port:
        return port
    raise ValueError("A source is required. Provide --source or --port.")


def open_source(source: str, baud: int, settle_s: float) -> SnapshotSource:
    if source.lower().startswith("tcp://"):
        parsed = urlparse(source)
        if not parsed.hostname or parsed.port is None:
            raise ValueError(f"Invalid TCP source: {source}")
        return TcpSnapshotSource(parsed.hostname, parsed.port, settle_s)
    if serial is None:
        raise RuntimeError("pyserial is required for serial sources. Install with: py -m pip install pyserial")
    return SerialSnapshotSource(source, baud, settle_s)


# Backward-compatible alias used by scan_mlx_z_stack.py.
def open_serial_port(port: str, baud: int, settle_s: float) -> SnapshotSource:
    return open_source(port, baud, settle_s)


def sensor_index_1based(tca1_ch: int, tca2_ch: int, addr: int) -> int:
    return tca1_ch * 16 + tca2_ch * 2 + (0 if addr == 0x60 else 1) + 1


def decompose_index_1based(index_1based: int) -> Tuple[int, int, int]:
    if not 1 <= index_1based <= 112:
        raise ValueError(f"Index out of range: {index_1based}")
    idx0 = index_1based - 1
    tca1_ch = idx0 // 16
    rem = idx0 % 16
    tca2_ch = rem // 2
    addr = 0x60 if (rem % 2) == 0 else 0x61
    return tca1_ch, tca2_ch, addr


def load_position_map(xlsx_path: pathlib.Path) -> Dict[int, Dict[str, object]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    required = ["Index", "Ring", "Ring Position", "Angle (°)", "r (mm)", "X (mm)", "Y (mm)"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(f"Mapping workbook is missing expected columns: {', '.join(missing)}")
    col = {name: headers.index(name) + 1 for name in headers}
    mapping: Dict[int, Dict[str, object]] = {}
    for row in range(2, ws.max_row + 1):
        idx = ws.cell(row, col["Index"]).value
        if idx is None:
            continue
        idx = int(idx)
        mapping[idx] = {
            "ring": ws.cell(row, col["Ring"]).value,
            "ring_position": ws.cell(row, col["Ring Position"]).value,
            "angle_deg": ws.cell(row, col["Angle (°)"]).value,
            "r_mm": ws.cell(row, col["r (mm)"]).value,
            "x_mm": ws.cell(row, col["X (mm)"]).value,
            "y_mm": ws.cell(row, col["Y (mm)"]).value,
        }
    if len(mapping) != 112:
        raise ValueError(f"Expected 112 index rows in mapping workbook, found {len(mapping)}")
    return mapping


def parse_sensor_cell(cell_text: str) -> Dict[str, object]:
    text = cell_text.strip()
    out = {
        "sensor_present": "",
        "measurement_available": "",
        "status": "",
        "x_uT": "",
        "y_uT": "",
        "z_uT": "",
        "temp_C": "",
        "age_ms": "",
        "flags": "",
    }
    if not text:
        out["status"] = "EMPTY"
        return out
    if text in STATUS_ONLY:
        out["status"] = text
        out["sensor_present"] = 0 if text in {"NOTCA2", "ABSENT"} else 1
        out["measurement_available"] = 0
        return out
    tokens = text.split()
    if len(tokens) < 6:
        out["status"] = f"UNPARSED:{text}"
        return out
    try:
        out.update(
            {
                "sensor_present": 1,
                "measurement_available": 1,
                "status": "OK",
                "x_uT": float(tokens[0]),
                "y_uT": float(tokens[1]),
                "z_uT": float(tokens[2]),
                "temp_C": float(tokens[3]),
                "age_ms": int(tokens[4]),
                "flags": " ".join(tokens[5:]),
            }
        )
        return out
    except ValueError:
        out["status"] = f"UNPARSED:{text}"
        return out


RowKey = Tuple[int, int, int]
ParsedSnapshot = Tuple[int, Dict[RowKey, Dict[str, object]], List[str]]


def parse_snapshot_text(snapshot_text: str) -> ParsedSnapshot:
    lines = [ln.rstrip("\r") for ln in snapshot_text.splitlines() if ln.strip()]
    if not lines:
        raise ValueError("Snapshot text was empty")
    snapshot_arduino_ms = None
    sensor_rows: Dict[RowKey, Dict[str, object]] = {}
    current_tca1 = None
    for line in lines:
        m = re.search(r"Snapshot @\s+(\d+)\s+ms", line)
        if m:
            snapshot_arduino_ms = int(m.group(1))
            continue
        m = re.match(r"TCA1 ch(\d+)\s+->", line)
        if m:
            current_tca1 = int(m.group(1))
            continue
        if current_tca1 is None:
            continue
        m = re.match(r"\s*(\d+)\s*\|\s*0x60\s*\|(.*)\|\s*0x61\s*\|(.*)$", line)
        if not m:
            continue
        tca2_ch = int(m.group(1))
        sensor_rows[(current_tca1, tca2_ch, 0x60)] = parse_sensor_cell(m.group(2))
        sensor_rows[(current_tca1, tca2_ch, 0x61)] = parse_sensor_cell(m.group(3))
    if snapshot_arduino_ms is None:
        raise ValueError("Could not find 'Snapshot @ ... ms' line in source data")
    return snapshot_arduino_ms, sensor_rows, lines


def send_serial_commands(
    source: SnapshotSource,
    commands: Iterable[str],
    inter_command_delay_s: float,
    echo: bool = True,
) -> None:
    for raw_cmd in commands:
        cmd = raw_cmd.strip()
        if not cmd:
            continue
        payload = (cmd + "\n").encode("utf-8")
        written = source.write(payload)
        source.flush()
        if written != len(payload):
            raise SerialCommandError(f"Failed to send full command: {cmd!r}")
        if echo:
            print(f">>> sent: {cmd}")
        if inter_command_delay_s > 0:
            time.sleep(inter_command_delay_s)


def drain_serial_lines(
    source: SnapshotSource,
    duration_s: float,
    line_limit: int = 500,
) -> List[str]:
    if duration_s <= 0:
        return []
    deadline = time.monotonic() + duration_s
    lines: List[str] = []
    while time.monotonic() < deadline and len(lines) < line_limit:
        raw = source.readline()
        if not raw:
            continue
        lines.append(raw.decode("utf-8", errors="replace").rstrip("\r\n"))
    return lines


def capture_one_snapshot_from_serial(
    source: SnapshotSource,
    timeout_s: float,
    raw_log_path: pathlib.Path | None,
) -> ParsedSnapshot:
    deadline = time.monotonic() + timeout_s
    collected_lines: List[str] = []
    inside_snapshot = False
    saw_header = False
    while time.monotonic() < deadline:
        raw = source.readline()
        if not raw:
            continue
        line = raw.decode("utf-8", errors="replace").rstrip("\r\n")
        if not inside_snapshot:
            if line.startswith(START_END_MARKER_PREFIX):
                inside_snapshot = True
                saw_header = False
                collected_lines = [line]
            continue
        collected_lines.append(line)
        if "Snapshot @" in line:
            saw_header = True
            continue
        if saw_header and line.startswith(START_END_MARKER_PREFIX):
            snapshot_text = "\n".join(collected_lines)
            if raw_log_path is not None:
                raw_log_path.write_text(snapshot_text + "\n", encoding="utf-8")
            return parse_snapshot_text(snapshot_text)
    raise TimeoutError(f"Timed out waiting for a complete snapshot on {source.source_name} after {timeout_s:.1f} s")


def capture_nth_snapshot_from_serial(
    source: SnapshotSource,
    timeout_s: float,
    raw_log_path: pathlib.Path | None = None,
    skip_snapshots: int = 0,
) -> ParsedSnapshot:
    last: ParsedSnapshot | None = None
    for i in range(skip_snapshots + 1):
        last = capture_one_snapshot_from_serial(
            source=source,
            timeout_s=timeout_s,
            raw_log_path=raw_log_path if i == skip_snapshots else None,
        )
    assert last is not None
    return last


def capture_one_snapshot(
    source: str,
    baud: int,
    timeout_s: float,
    settle_s: float,
    raw_log_path: pathlib.Path | None,
    pre_commands: List[str] | None = None,
    post_commands: List[str] | None = None,
    inter_command_delay_s: float = 0.2,
    pre_capture_delay_s: float = 0.0,
    post_command_read_s: float = 0.0,
) -> ParsedSnapshot:
    pre_commands = pre_commands or []
    post_commands = post_commands or []
    source_obj = open_source(source=source, baud=baud, settle_s=settle_s)
    try:
        if pre_commands:
            send_serial_commands(source_obj, pre_commands, inter_command_delay_s=inter_command_delay_s)
        if pre_capture_delay_s > 0:
            time.sleep(pre_capture_delay_s)
        snapshot = capture_one_snapshot_from_serial(source_obj, timeout_s=timeout_s, raw_log_path=raw_log_path)
        if post_commands:
            send_serial_commands(source_obj, post_commands, inter_command_delay_s=inter_command_delay_s)
            drained = drain_serial_lines(source_obj, duration_s=post_command_read_s)
            if drained:
                print("Source after capture:")
                for line in drained:
                    print(f"  {line}")
        return snapshot
    finally:
        source_obj.close()


def build_output_rows(
    mapping: Dict[int, Dict[str, object]],
    snapshot_arduino_ms: int,
    capture_timestamp: dt.datetime,
    parsed_rows: Dict[RowKey, Dict[str, object]],
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    capture_iso = capture_timestamp.isoformat(timespec="seconds")
    for idx in range(1, 113):
        tca1_ch, tca2_ch, addr = decompose_index_1based(idx)
        mapping_row = mapping[idx]
        parsed = parsed_rows.get((tca1_ch, tca2_ch, addr), None)
        row = {
            "index": idx,
            "tca1_channel": tca1_ch,
            "tca2_channel": tca2_ch,
            "mlx_address": f"0x{addr:02X}",
            "ring": mapping_row["ring"],
            "ring_position": mapping_row["ring_position"],
            "angle_deg": mapping_row["angle_deg"],
            "r_mm": mapping_row["r_mm"],
            "x_mm": mapping_row["x_mm"],
            "y_mm": mapping_row["y_mm"],
            "sensor_present": 0,
            "measurement_available": 0,
            "status": "ABSENT",
            "x_uT": "",
            "y_uT": "",
            "z_uT": "",
            "temp_C": "",
            "age_ms": "",
            "flags": "",
            "snapshot_arduino_ms": snapshot_arduino_ms,
            "capture_pc_local_time": capture_iso,
        }
        if parsed is not None:
            row.update(parsed)
        rows.append(row)
    return rows


def build_vector_points(
    mapping: Dict[int, Dict[str, object]],
    parsed_rows: Dict[RowKey, Dict[str, object]],
    z_mm: float,
) -> List[Dict[str, object]]:
    points: List[Dict[str, object]] = []
    for idx in range(1, 113):
        tca1_ch, tca2_ch, addr = decompose_index_1based(idx)
        parsed = parsed_rows.get((tca1_ch, tca2_ch, addr))
        if parsed is None or parsed.get("measurement_available") != 1:
            continue
        try:
            x_mm = float(mapping[idx]["x_mm"])
            y_mm = float(mapping[idx]["y_mm"])
            bx_uT = float(parsed["x_uT"])
            by_uT = float(parsed["y_uT"])
            bz_uT = float(parsed["z_uT"])
            temp_C = float(parsed["temp_C"])
            age_ms = int(parsed["age_ms"])
        except (TypeError, ValueError, KeyError):
            continue
        magnitude_uT = (bx_uT ** 2 + by_uT ** 2 + bz_uT ** 2) ** 0.5
        points.append(
            {
                "index": idx,
                "tca1_channel": tca1_ch,
                "tca2_channel": tca2_ch,
                "mlx_address": f"0x{addr:02X}",
                "x_mm": x_mm,
                "y_mm": y_mm,
                "z_mm": float(z_mm),
                "bx_uT": bx_uT,
                "by_uT": by_uT,
                "bz_uT": bz_uT,
                "magnitude_uT": magnitude_uT,
                "temp_C": temp_C,
                "age_ms": age_ms,
                "flags": str(parsed.get("flags", "")),
            }
        )
    return points


def _sensor_wide_prefix(idx: int) -> str:
    return f"i{idx:03d}_"


def build_wide_snapshot_row(
    mapping: Dict[int, Dict[str, object]],
    snapshot_arduino_ms: int,
    capture_timestamp: dt.datetime,
    parsed_rows: Dict[RowKey, Dict[str, object]],
    *,
    z_index: int,
    z_steps: int,
    z_um: float,
    z_mm: float,
) -> Dict[str, object]:
    capture_iso = capture_timestamp.isoformat(timespec="seconds")
    row: Dict[str, object] = {
        "z_index": z_index,
        "z_steps": z_steps,
        "z_um": z_um,
        "z_mm": z_mm,
        "snapshot_arduino_ms": snapshot_arduino_ms,
        "capture_pc_local_time": capture_iso,
        "sensors_present": 0,
        "sensors_measured": 0,
    }
    present_count = 0
    measured_count = 0
    for idx in range(1, 113):
        tca1_ch, tca2_ch, addr = decompose_index_1based(idx)
        mapping_row = mapping[idx]
        parsed = parsed_rows.get((tca1_ch, tca2_ch, addr), None)
        prefix = _sensor_wide_prefix(idx)
        row[prefix + "index"] = idx
        row[prefix + "tca1_channel"] = tca1_ch
        row[prefix + "tca2_channel"] = tca2_ch
        row[prefix + "mlx_address"] = f"0x{addr:02X}"
        row[prefix + "ring"] = mapping_row["ring"]
        row[prefix + "ring_position"] = mapping_row["ring_position"]
        row[prefix + "angle_deg"] = mapping_row["angle_deg"]
        row[prefix + "r_mm"] = mapping_row["r_mm"]
        row[prefix + "x_mm"] = mapping_row["x_mm"]
        row[prefix + "y_mm"] = mapping_row["y_mm"]
        row[prefix + "sensor_present"] = 0
        row[prefix + "measurement_available"] = 0
        row[prefix + "status"] = "ABSENT"
        row[prefix + "x_uT"] = ""
        row[prefix + "y_uT"] = ""
        row[prefix + "z_uT"] = ""
        row[prefix + "temp_C"] = ""
        row[prefix + "age_ms"] = ""
        row[prefix + "flags"] = ""
        if parsed is not None:
            row[prefix + "sensor_present"] = parsed.get("sensor_present", 0)
            row[prefix + "measurement_available"] = parsed.get("measurement_available", 0)
            row[prefix + "status"] = parsed.get("status", "")
            row[prefix + "x_uT"] = parsed.get("x_uT", "")
            row[prefix + "y_uT"] = parsed.get("y_uT", "")
            row[prefix + "z_uT"] = parsed.get("z_uT", "")
            row[prefix + "temp_C"] = parsed.get("temp_C", "")
            row[prefix + "age_ms"] = parsed.get("age_ms", "")
            row[prefix + "flags"] = parsed.get("flags", "")
        present_count += 1 if row[prefix + "sensor_present"] == 1 else 0
        measured_count += 1 if row[prefix + "measurement_available"] == 1 else 0
    row["sensors_present"] = present_count
    row["sensors_measured"] = measured_count
    return row


def wide_snapshot_fieldnames() -> List[str]:
    fields = list(WIDE_BASE_FIELDS)
    for idx in range(1, 113):
        prefix = _sensor_wide_prefix(idx)
        fields.extend(
            [
                prefix + "index",
                prefix + "tca1_channel",
                prefix + "tca2_channel",
                prefix + "mlx_address",
                prefix + "ring",
                prefix + "ring_position",
                prefix + "angle_deg",
                prefix + "r_mm",
                prefix + "x_mm",
                prefix + "y_mm",
                prefix + "sensor_present",
                prefix + "measurement_available",
                prefix + "status",
                prefix + "x_uT",
                prefix + "y_uT",
                prefix + "z_uT",
                prefix + "temp_C",
                prefix + "age_ms",
                prefix + "flags",
            ]
        )
    return fields


def append_wide_snapshot_csv_row(out_path: pathlib.Path, row: Dict[str, object]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = wide_snapshot_fieldnames()
    write_header = not out_path.exists() or out_path.stat().st_size == 0
    with out_path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def write_csv(rows: List[Dict[str, object]], out_path: pathlib.Path) -> None:
    fieldnames = [
        "index",
        "tca1_channel",
        "tca2_channel",
        "mlx_address",
        "ring",
        "ring_position",
        "angle_deg",
        "r_mm",
        "x_mm",
        "y_mm",
        "sensor_present",
        "measurement_available",
        "status",
        "x_uT",
        "y_uT",
        "z_uT",
        "temp_C",
        "age_ms",
        "flags",
        "snapshot_arduino_ms",
        "capture_pc_local_time",
    ]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def default_output_path() -> pathlib.Path:
    stamp = dt.datetime.now().strftime("mlx_snapshot_%Y%m%d_%H%M%S.csv")
    return SCRIPT_DIR / stamp


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Capture one MLX90394 snapshot from a serial or TCP source and save it to CSV."
    )
    p.add_argument("--source", help="Snapshot source: COM port like COM7, or tcp://127.0.0.1:8765")
    p.add_argument("--port", help="Backward-compatible alias for --source when using serial COM ports")
    p.add_argument("--baud", type=int, default=115200, help="Baud rate for serial sources (default: 115200)")
    p.add_argument(
        "--mapping",
        type=pathlib.Path,
        default=SCRIPT_DIR / "SensorGrid_112_Positions (1).xlsx",
        help="Path to the mapping workbook (.xlsx)",
    )
    p.add_argument("--output", type=pathlib.Path, default=None, help="Output CSV path")
    p.add_argument("--raw-log", type=pathlib.Path, default=None, help="Optional path to save raw snapshot text")
    p.add_argument("--timeout", type=float, default=20.0, help="Seconds to wait for a complete snapshot")
    p.add_argument("--settle", type=float, default=2.5, help="Seconds to wait after opening the source")
    p.add_argument("--list-ports", action="store_true", help="List available serial ports and exit")
    p.add_argument("--command-only", action="store_true", help="Send command(s) and exit without capturing")
    p.add_argument("--send", action="append", default=[], help="Send an arbitrary command line before capture")
    p.add_argument("--start", action="store_true", help="Send 'start' before capture or before exiting")
    p.add_argument("--stop", action="store_true", help="Send 'stop' before capture or before exiting")
    p.add_argument("--scan", action="store_true", help="Send 'scan' before capture or before exiting")
    p.add_argument("--stop-after-capture", action="store_true", help="Send 'stop' after a successful capture")
    p.add_argument("--command-delay", type=float, default=0.2, help="Delay in seconds between commands")
    p.add_argument("--pre-capture-delay", type=float, default=0.5, help="Delay after pre-capture commands")
    p.add_argument("--read-after-command", type=float, default=1.0, help="How long to read source output after commands")
    return p.parse_args()


def print_ports() -> None:
    if list_ports is None:
        print("pyserial is required to list COM ports. Install with: py -m pip install pyserial", file=sys.stderr)
        return
    ports = list(list_ports.comports())
    if not ports:
        print("No serial ports found.")
        return
    print("Available serial ports:")
    for p in ports:
        desc = f" - {p.description}" if p.description else ""
        print(f"  {p.device}{desc}")


def command_list_from_args(args: argparse.Namespace) -> List[str]:
    cmds: List[str] = []
    if args.start:
        cmds.append("start")
    if args.stop:
        cmds.append("stop")
    if args.scan:
        cmds.append("scan")
    cmds.extend(args.send)
    return cmds


def run_command_only(
    source_name: str,
    baud: int,
    settle_s: float,
    commands: List[str],
    command_delay_s: float,
    read_after_command_s: float,
) -> int:
    if not commands:
        print("Error: --command-only requires at least one command.", file=sys.stderr)
        return 2
    source = open_source(source_name, baud=baud, settle_s=settle_s)
    try:
        send_serial_commands(source, commands, inter_command_delay_s=command_delay_s)
        response_lines = drain_serial_lines(source, duration_s=read_after_command_s)
    finally:
        source.close()
    if response_lines:
        print("Source response:")
        for line in response_lines:
            print(f"  {line}")
    else:
        print("No source response received within the read window.")
    return 0


def main() -> int:
    args = parse_args()
    if args.list_ports:
        print_ports()
        return 0
    try:
        source_name = normalize_source(args.source, args.port)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2

    pre_commands = command_list_from_args(args)
    if args.command_only:
        try:
            return run_command_only(
                source_name=source_name,
                baud=args.baud,
                settle_s=args.settle,
                commands=pre_commands,
                command_delay_s=args.command_delay,
                read_after_command_s=args.read_after_command,
            )
        except Exception as exc:
            print(f"Error: {exc}", file=sys.stderr)
            return 1

    mapping_path = args.mapping.resolve()
    if not mapping_path.exists():
        print(f"Error: mapping workbook not found: {mapping_path}", file=sys.stderr)
        return 2
    output_path = args.output.resolve() if args.output else default_output_path().resolve()
    if output_path.suffix.lower() != ".csv":
        output_path = output_path.with_suffix(".csv")
    raw_log_path = args.raw_log.resolve() if args.raw_log else None
    post_commands = ["stop"] if args.stop_after_capture else []
    try:
        mapping = load_position_map(mapping_path)
        snapshot_arduino_ms, parsed_rows, _lines = capture_one_snapshot(
            source=source_name,
            baud=args.baud,
            timeout_s=args.timeout,
            settle_s=args.settle,
            raw_log_path=raw_log_path,
            pre_commands=pre_commands,
            post_commands=post_commands,
            inter_command_delay_s=args.command_delay,
            pre_capture_delay_s=args.pre_capture_delay,
            post_command_read_s=args.read_after_command,
        )
        capture_ts = dt.datetime.now().astimezone()
        rows = build_output_rows(mapping, snapshot_arduino_ms, capture_ts, parsed_rows)
        write_csv(rows, output_path)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    present_count = sum(1 for r in rows if r["sensor_present"] == 1)
    measured_count = sum(1 for r in rows if r["measurement_available"] == 1)
    print(f"Saved snapshot CSV: {output_path}")
    print(f"Snapshot source: {source_name}")
    print(f"Snapshot Arduino/source time: {snapshot_arduino_ms} ms")
    print(f"Sensors present: {present_count} / 112")
    print(f"Sensors with numeric measurements: {measured_count} / 112")
    if raw_log_path is not None:
        print(f"Saved raw snapshot text: {raw_log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
