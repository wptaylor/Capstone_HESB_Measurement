#!/usr/bin/env python3
"""
Run an MLX90394 z-stack scan, append one wide CSV row per z position,
and visualize the field live in a localhost browser viewer.

Real mode:
- moves the z stage
- restarts MLX acquisition after each move
- discards the first snapshot after restart
- captures the next complete settled snapshot

Dummy mode:
- does not touch the motor
- talks to dummy_mlx_source.py over TCP by default
- advances z logically in software

The browser viewer is used instead of repeated 3D pyplot redraws because it stays
interactive while the scan is running and scales much better for thousands of 3D vectors.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import pathlib
import socket
import sys
import threading
import time
import webbrowser
from dataclasses import dataclass
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from typing import Dict, List, Optional, Sequence, Tuple
from urllib.parse import urlparse

import openpyxl

try:
    import serial  # type: ignore
except Exception:  # pragma: no cover
    serial = None

SCRIPT_DIR = pathlib.Path(__file__).resolve().parent
DEFAULT_STATE_PATH = SCRIPT_DIR / "z_stage_state.json"
DEFAULT_MAPPING_XLSX = SCRIPT_DIR / "SensorGrid_112_Positions (1).xlsx"
DEFAULT_OUTPUT_CSV = SCRIPT_DIR / f"mlx_z_stack_{dt.datetime.now():%Y%m%d_%H%M%S}.csv"
DEFAULT_PROGRESS_JSON = SCRIPT_DIR / "mlx_z_stack_progress.json"
WEB_STATE_DIR = SCRIPT_DIR / ".mlx_live_view"

DEFAULT_BAUD = 115200
STEP_SIZE_UM = 1.5

MLX_SOURCE_DEFAULT = "COM7"
DUMMY_SOURCE_DEFAULT = "tcp://127.0.0.1:8765"
STAGE_PORT_DEFAULT = "COM8"

Z_START_MM_DEFAULT = 0.0
Z_STOP_MM_DEFAULT = 60.0
Z_STEP_MM_DEFAULT = 1.5

MLX_SERIAL_SETTLE_S = 1.0
STAGE_SERIAL_SETTLE_S = 2.5
STAGE_STARTUP_READ_S = 0.8
STAGE_MOVE_TIMEOUT_S = 180.0
STAGE_SETTLE_AFTER_MOVE_S = 0.25

MLX_COMMAND_DELAY_S = 0.12
MLX_POST_COMMAND_READ_S = 0.35
MLX_CAPTURE_TIMEOUT_S = 15.0
MLX_SKIP_SNAPSHOTS_AFTER_START = 1

PLOT_RADIUS_MM = 35.0
PLOT_CIRCLE_RADIUS_MM = 17.5
PLOT_ARROW_LENGTH_MM = 1.0
PLOT_MAG_MAX_UT = 2000.0
PLOT_LINE_WIDTH_PX = 4.0
PLOT_ORIGIN_MARKER_SIZE = 3.0
PLOT_TIP_MARKER_SIZE = 2.5
WEB_PORT_DEFAULT = 8766
WEB_REFRESH_MS = 400
COLOR_BINS = 24
TOTAL_SENSORS = 112


@dataclass
class StageStateSimple:
    current_steps: int = 0


@dataclass
class SensorMeasurement:
    sensor_present: int
    measurement_available: int
    status: str
    x_uT: object
    y_uT: object
    z_uT: object
    temp_C: object
    age_ms: object
    flags: str


@dataclass
class SnapshotData:
    snapshot_arduino_ms: int
    rows: Dict[Tuple[int, int, int], SensorMeasurement]
    raw_lines: List[str]


class SourceBase:
    def readline(self) -> bytes:
        raise NotImplementedError

    def write_line(self, text: str) -> None:
        raise NotImplementedError

    def close(self) -> None:
        raise NotImplementedError


class SerialSource(SourceBase):
    def __init__(self, port: str, baud: int, settle_s: float) -> None:
        if serial is None:
            raise RuntimeError("pyserial is required for serial sources. Install with: py -m pip install pyserial")
        self.ser = serial.Serial(port=port, baudrate=baud, timeout=0.25, write_timeout=2)
        time.sleep(max(0.0, settle_s))
        self.ser.reset_input_buffer()
        self.ser.reset_output_buffer()

    def readline(self) -> bytes:
        return self.ser.readline()

    def write_line(self, text: str) -> None:
        payload = (text.rstrip("\r\n") + "\n").encode("utf-8")
        written = self.ser.write(payload)
        self.ser.flush()
        if written != len(payload):
            raise IOError("failed to write full command to serial source")

    def close(self) -> None:
        self.ser.close()


class TCPSource(SourceBase):
    def __init__(self, host: str, port: int, settle_s: float) -> None:
        self.sock = socket.create_connection((host, port), timeout=5.0)
        self.sock.settimeout(0.25)
        self.buf = bytearray()
        if settle_s > 0:
            deadline = time.monotonic() + settle_s
            while time.monotonic() < deadline:
                try:
                    chunk = self.sock.recv(4096)
                except socket.timeout:
                    continue
                if not chunk:
                    break
                self.buf.extend(chunk)
        self.buf.clear()

    def readline(self) -> bytes:
        while True:
            if b"\n" in self.buf:
                line, _, rest = self.buf.partition(b"\n")
                self.buf = bytearray(rest)
                return bytes(line + b"\n")
            try:
                chunk = self.sock.recv(4096)
            except socket.timeout:
                return b""
            if not chunk:
                return b""
            self.buf.extend(chunk)

    def write_line(self, text: str) -> None:
        payload = (text.rstrip("\r\n") + "\n").encode("utf-8")
        self.sock.sendall(payload)

    def close(self) -> None:
        self.sock.close()


def open_source(source_name: str, baud: int, settle_s: float) -> SourceBase:
    if source_name.lower().startswith("tcp://"):
        parsed = urlparse(source_name)
        host = parsed.hostname or "127.0.0.1"
        if parsed.port is None:
            raise ValueError(f"TCP source is missing port: {source_name}")
        return TCPSource(host, parsed.port, settle_s)
    return SerialSource(source_name, baud, settle_s)


def drain_source_lines(source: SourceBase, duration_s: float) -> List[str]:
    deadline = time.monotonic() + max(0.0, duration_s)
    lines: List[str] = []
    while time.monotonic() < deadline:
        raw = source.readline()
        if not raw:
            continue
        lines.append(raw.decode("utf-8", errors="replace").rstrip("\r\n"))
    return lines


def send_commands(source: SourceBase, commands: Sequence[str], inter_command_delay_s: float = 0.1) -> None:
    for i, cmd in enumerate(commands):
        source.write_line(cmd)
        if i != len(commands) - 1:
            time.sleep(max(0.0, inter_command_delay_s))


def start_mlx_stream(source: SourceBase) -> None:
    send_commands(source, ["start"], inter_command_delay_s=MLX_COMMAND_DELAY_S)
    _ = drain_source_lines(source, MLX_POST_COMMAND_READ_S)


def stop_mlx_stream(source: SourceBase) -> None:
    send_commands(source, ["stop"], inter_command_delay_s=MLX_COMMAND_DELAY_S)
    _ = drain_source_lines(source, MLX_POST_COMMAND_READ_S)


def set_dummy_z(source: SourceBase, z_mm: float) -> None:
    send_commands(source, [f"setz {z_mm:.6f}"], inter_command_delay_s=MLX_COMMAND_DELAY_S)
    _ = drain_source_lines(source, MLX_POST_COMMAND_READ_S)


def atomic_write_text(path: pathlib.Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)


def write_json(path: pathlib.Path, payload: dict) -> None:
    atomic_write_text(path, json.dumps(payload, indent=2, sort_keys=True) + "\n")


def load_position_map(xlsx_path: pathlib.Path) -> Dict[int, Dict[str, object]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    required = ["Index", "Ring", "Ring Position", "Angle (°)", "r (mm)", "X (mm)", "Y (mm)"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(f"Mapping workbook is missing expected columns: {', '.join(missing)}")
    col = {name: headers.index(name) + 1 for name in headers}
    mapping: Dict[int, Dict[str, object]] = {}
    for row in range(2, ws.max_row + 1):
        idx = ws.cell(row, col["Index"]).value
        if idx is None:
            continue
        idx = int(idx)
        mapping[idx] = {
            "index": idx,
            "ring": ws.cell(row, col["Ring"]).value,
            "ring_position": ws.cell(row, col["Ring Position"]).value,
            "angle_deg": ws.cell(row, col["Angle (°)"]).value,
            "r_mm": ws.cell(row, col["r (mm)"]).value,
            "x_mm": ws.cell(row, col["X (mm)"]).value,
            "y_mm": ws.cell(row, col["Y (mm)"]).value,
        }
    if len(mapping) != TOTAL_SENSORS:
        raise ValueError(f"Expected {TOTAL_SENSORS} mapping rows, found {len(mapping)}")
    return mapping


def decompose_index_1based(index_1based: int) -> Tuple[int, int, int]:
    idx0 = index_1based - 1
    tca1_ch = idx0 // 16
    rem = idx0 % 16
    tca2_ch = rem // 2
    addr = 0x60 if (rem % 2) == 0 else 0x61
    return tca1_ch, tca2_ch, addr


STATUS_ONLY = {"NOTCA2", "ABSENT", "EMPTY", "NODATA"}


def parse_sensor_cell(text: str) -> SensorMeasurement:
    stripped = text.strip()
    if not stripped:
        return SensorMeasurement(0, 0, "EMPTY", "", "", "", "", "", "")
    if stripped in STATUS_ONLY:
        present = 0 if stripped in {"NOTCA2", "ABSENT"} else 1
        return SensorMeasurement(present, 0, stripped, "", "", "", "", "", "")
    toks = stripped.split()
    if len(toks) < 6:
        return SensorMeasurement(1, 0, f"UNPARSED:{stripped}", "", "", "", "", "", "")
    try:
        x_uT = float(toks[0])
        y_uT = float(toks[1])
        z_uT = float(toks[2])
        temp_C = float(toks[3])
        age_ms = int(float(toks[4]))
        flags = toks[5]
    except Exception:
        return SensorMeasurement(1, 0, f"UNPARSED:{stripped}", "", "", "", "", "", "")
    return SensorMeasurement(1, 1, "OK", x_uT, y_uT, z_uT, temp_C, age_ms, flags)


def parse_snapshot_text(snapshot_text: str) -> SnapshotData:
    lines = [line.rstrip("\r\n") for line in snapshot_text.splitlines()]
    snapshot_ms: Optional[int] = None
    current_tca1: Optional[int] = None
    rows: Dict[Tuple[int, int, int], SensorMeasurement] = {}

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("Snapshot @"):
            try:
                snapshot_ms = int(line.split("@", 1)[1].split("ms", 1)[0].strip())
            except Exception as exc:
                raise ValueError(f"Failed to parse snapshot timestamp from line: {raw_line}") from exc
            continue
        if line.startswith("TCA1 ch"):
            try:
                current_tca1 = int(line.split("TCA1 ch", 1)[1].split()[0])
            except Exception:
                current_tca1 = None
            continue
        if current_tca1 is None:
            continue
        if "| 0x60 |" in raw_line and "| 0x61 |" in raw_line:
            parts = raw_line.split("|")
            if len(parts) < 5:
                continue
            try:
                tca2_ch = int(parts[0].strip())
            except Exception:
                continue
            cell60 = parts[2]
            cell61 = parts[4]
            rows[(current_tca1, tca2_ch, 0x60)] = parse_sensor_cell(cell60)
            rows[(current_tca1, tca2_ch, 0x61)] = parse_sensor_cell(cell61)

    if snapshot_ms is None:
        raise ValueError("No snapshot timestamp found in snapshot text")
    return SnapshotData(snapshot_ms, rows, lines)


def capture_nth_snapshot(source: SourceBase, timeout_s: float, skip_snapshots: int = 0, raw_log_path: Optional[pathlib.Path] = None) -> SnapshotData:
    deadline = time.monotonic() + timeout_s
    current: Optional[List[str]] = None
    skipped = 0
    while time.monotonic() < deadline:
        raw = source.readline()
        if not raw:
            continue
        line = raw.decode("utf-8", errors="replace").rstrip("\r\n")
        if current is None:
            if line.strip().startswith("Snapshot @"):
                current = [line]
            continue
        current.append(line)
        if line.strip().startswith("###"):
            snapshot_text = "\n".join(current)
            current = None
            snapshot = parse_snapshot_text(snapshot_text)
            if skipped < skip_snapshots:
                skipped += 1
                continue
            if raw_log_path is not None:
                raw_log_path.write_text(snapshot_text + "\n", encoding="utf-8")
            return snapshot
    raise TimeoutError(f"Timed out waiting for a complete snapshot after {timeout_s:.1f} s")


def capture_settled_snapshot(source: SourceBase, timeout_s: float, raw_log_path: Optional[pathlib.Path]) -> SnapshotData:
    start_mlx_stream(source)
    try:
        return capture_nth_snapshot(source, timeout_s=timeout_s, skip_snapshots=MLX_SKIP_SNAPSHOTS_AFTER_START, raw_log_path=raw_log_path)
    finally:
        stop_mlx_stream(source)


def load_stage_state(path: pathlib.Path) -> StageStateSimple:
    if not path.exists():
        return StageStateSimple(0)
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return StageStateSimple(int(data.get("current_steps", 0)))
    except Exception:
        return StageStateSimple(0)


def steps_from_target_um_local(target_um: float) -> int:
    return int(round(float(target_um) / STEP_SIZE_UM))


def target_um_from_steps_local(steps: int) -> float:
    return int(steps) * STEP_SIZE_UM


def steps_from_mm_exact(mm: float, label: str) -> int:
    steps = steps_from_target_um_local(mm * 1000.0)
    actual_mm = target_um_from_steps_local(steps) / 1000.0
    if abs(actual_mm - mm) > 1e-9:
        raise ValueError(
            f"{label}={mm} mm does not fall on an integer motor step. "
            f"Nearest realizable value is {actual_mm:.6f} mm ({steps} steps)."
        )
    return steps


def build_z_targets_steps(start_mm: float, stop_mm: float, step_mm: float) -> List[int]:
    if step_mm <= 0:
        raise ValueError("step_mm must be > 0")
    start_steps = steps_from_mm_exact(start_mm, "start_mm")
    stop_steps = steps_from_mm_exact(stop_mm, "stop_mm")
    step_steps = steps_from_mm_exact(step_mm, "step_mm")
    if stop_steps < start_steps:
        raise ValueError("stop_mm must be >= start_mm")
    if ((stop_steps - start_steps) % step_steps) != 0:
        raise ValueError("The z range is not an integer multiple of the step size")
    return list(range(start_steps, stop_steps + 1, step_steps))


def load_stage_api() -> dict:
    try:
        import set_z_height as szh
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Real stage control requires set_z_height.py in the same folder as scan_mlx_z_stack.py. "
            "For dummy mode, use --dummy-data."
        ) from exc
    return {
        "load_state": szh.load_state,
        "open_stage_serial_port": szh.open_serial_port,
        "save_state": szh.save_state,
        "send_relative_steps_and_wait": szh.send_relative_steps_and_wait,
        "drain_stage_serial": szh.drain_serial,
        "stage_direction_sign": int(szh.STAGE_DIRECTION_SIGN),
    }


def move_stage_absolute(ser, state_path: pathlib.Path, target_steps: int, timeout_s: float) -> Tuple[int, float]:
    stage_api = load_stage_api()
    state = stage_api["load_state"](state_path)
    delta_steps = target_steps - state.current_steps
    if delta_steps == 0:
        return 0, 0.0
    move_result = stage_api["send_relative_steps_and_wait"](ser, delta_steps, timeout_s)
    expected_device_delta = stage_api["stage_direction_sign"] * delta_steps
    if move_result.confirmed_delta_steps != expected_device_delta:
        raise RuntimeError(
            f"Stage controller confirmed {move_result.confirmed_delta_steps} steps, expected {delta_steps}"
        )
    state.current_steps += delta_steps
    stage_api["save_state"](state_path, state, note=f"scan_mlx_z_stack moved by {delta_steps} steps")
    return delta_steps, float(move_result.duration_s)


def build_vector_points(mapping: Dict[int, Dict[str, object]], snapshot_rows: Dict[Tuple[int, int, int], SensorMeasurement], z_mm: float) -> List[Dict[str, object]]:
    points: List[Dict[str, object]] = []
    for idx in range(1, TOTAL_SENSORS + 1):
        tca1, tca2, addr = decompose_index_1based(idx)
        meas = snapshot_rows.get((tca1, tca2, addr))
        if meas is None or meas.measurement_available != 1:
            continue
        points.append({
            "index": idx,
            "tca1_channel": tca1,
            "tca2_channel": tca2,
            "mlx_address": f"0x{addr:02X}",
            "x_mm": float(mapping[idx]["x_mm"]),
            "y_mm": float(mapping[idx]["y_mm"]),
            "z_mm": float(z_mm),
            "bx_uT": float(meas.x_uT),
            "by_uT": float(meas.y_uT),
            "bz_uT": float(meas.z_uT),
            "temp_C": float(meas.temp_C),
            "age_ms": int(meas.age_ms),
            "flags": meas.flags,
            "status": meas.status,
        })
    return points


def csv_fieldnames() -> List[str]:
    fields = [
        "z_index",
        "z_steps",
        "z_um",
        "z_mm",
        "snapshot_arduino_ms",
        "capture_pc_local_time",
        "sensors_present",
        "sensors_measured",
    ]
    for idx in range(1, TOTAL_SENSORS + 1):
        fields.extend([
            f"s{idx:03d}_index",
            f"s{idx:03d}_tca1_channel",
            f"s{idx:03d}_tca2_channel",
            f"s{idx:03d}_mlx_address",
            f"s{idx:03d}_ring",
            f"s{idx:03d}_ring_position",
            f"s{idx:03d}_angle_deg",
            f"s{idx:03d}_r_mm",
            f"s{idx:03d}_x_mm",
            f"s{idx:03d}_y_mm",
            f"s{idx:03d}_sensor_present",
            f"s{idx:03d}_measurement_available",
            f"s{idx:03d}_status",
            f"s{idx:03d}_x_uT",
            f"s{idx:03d}_y_uT",
            f"s{idx:03d}_z_uT",
            f"s{idx:03d}_temp_C",
            f"s{idx:03d}_age_ms",
            f"s{idx:03d}_flags",
        ])
    return fields


def build_wide_row(mapping: Dict[int, Dict[str, object]], snapshot: SnapshotData, capture_ts: dt.datetime, z_index: int, z_steps: int, z_um: float, z_mm: float) -> Dict[str, object]:
    row: Dict[str, object] = {
        "z_index": z_index,
        "z_steps": z_steps,
        "z_um": z_um,
        "z_mm": z_mm,
        "snapshot_arduino_ms": snapshot.snapshot_arduino_ms,
        "capture_pc_local_time": capture_ts.isoformat(timespec="seconds"),
        "sensors_present": 0,
        "sensors_measured": 0,
    }
    present = 0
    measured = 0
    for idx in range(1, TOTAL_SENSORS + 1):
        tca1, tca2, addr = decompose_index_1based(idx)
        mapping_row = mapping[idx]
        meas = snapshot.rows.get((tca1, tca2, addr), SensorMeasurement(0, 0, "ABSENT", "", "", "", "", "", ""))
        present += 1 if meas.sensor_present == 1 else 0
        measured += 1 if meas.measurement_available == 1 else 0
        row.update({
            f"s{idx:03d}_index": idx,
            f"s{idx:03d}_tca1_channel": tca1,
            f"s{idx:03d}_tca2_channel": tca2,
            f"s{idx:03d}_mlx_address": f"0x{addr:02X}",
            f"s{idx:03d}_ring": mapping_row["ring"],
            f"s{idx:03d}_ring_position": mapping_row["ring_position"],
            f"s{idx:03d}_angle_deg": mapping_row["angle_deg"],
            f"s{idx:03d}_r_mm": mapping_row["r_mm"],
            f"s{idx:03d}_x_mm": mapping_row["x_mm"],
            f"s{idx:03d}_y_mm": mapping_row["y_mm"],
            f"s{idx:03d}_sensor_present": meas.sensor_present,
            f"s{idx:03d}_measurement_available": meas.measurement_available,
            f"s{idx:03d}_status": meas.status,
            f"s{idx:03d}_x_uT": meas.x_uT,
            f"s{idx:03d}_y_uT": meas.y_uT,
            f"s{idx:03d}_z_uT": meas.z_uT,
            f"s{idx:03d}_temp_C": meas.temp_C,
            f"s{idx:03d}_age_ms": meas.age_ms,
            f"s{idx:03d}_flags": meas.flags,
        })
    row["sensors_present"] = present
    row["sensors_measured"] = measured
    return row


def append_wide_csv(path: pathlib.Path, fieldnames: Sequence[str], row: Dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists() or path.stat().st_size == 0
    with path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        writer.writerow(row)


class SilentLogHandler(SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return


INDEX_HTML = f"""<!doctype html>
<html>
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <title>MLX Live Field Map</title>
  <script src=\"https://cdn.plot.ly/plotly-2.35.2.min.js\"></script>
  <style>
    html, body {{ margin:0; padding:0; width:100%; height:100%; background:#111; color:#ddd; font-family:Arial, sans-serif; overflow:hidden; }}
    #plot {{ width:100vw; height:100vh; }}
    #status {{ position:fixed; top:8px; left:12px; z-index:10; color:#ddd; font-size:14px; background:rgba(0,0,0,0.25); padding:4px 8px; border-radius:6px; }}
  </style>
</head>
<body>
<div id=\"status\">Connecting…</div>
<div id=\"plot\"></div>
<script>
const plotDiv = document.getElementById('plot');
const statusDiv = document.getElementById('status');
let firstLoad = true;
async function refresh() {{
  try {{
    const resp = await fetch('state.json?ts=' + Date.now(), {{cache:'no-store'}});
    const state = await resp.json();
    statusDiv.textContent = state.status || '';
    const layout = state.layout || {{}};
    layout.uirevision = 'keep-camera';
    const config = {{responsive:true, scrollZoom:true, displaylogo:false}};
    if (firstLoad) {{
      await Plotly.newPlot(plotDiv, state.data || [], layout, config);
      firstLoad = false;
    }} else {{
      await Plotly.react(plotDiv, state.data || [], layout, config);
    }}
  }} catch (err) {{
    statusDiv.textContent = 'Viewer waiting for data…';
  }}
}}
setInterval(refresh, {WEB_REFRESH_MS});
refresh();
</script>
</body>
</html>
"""


def reference_circle_xyz(radius_mm: float, n: int = 361) -> Tuple[List[float], List[float], List[float]]:
    xs: List[float] = []
    ys: List[float] = []
    zs: List[float] = []
    for i in range(n):
        theta = 2.0 * math.pi * i / (n - 1)
        xs.append(radius_mm * math.cos(theta))
        ys.append(radius_mm * math.sin(theta))
        zs.append(0.0)
    return xs, ys, zs


def color_for_fraction(frac: float) -> str:
    frac = max(0.0, min(1.0, frac))
    r = int(round(255.0 * frac))
    g = 0
    b = int(round(255.0 * (1.0 - frac)))
    return f"rgb({r},{g},{b})"


def build_plotly_payload(all_points: Sequence[Dict[str, object]], title: str, status: str) -> dict:
    origin_x: List[float] = []
    origin_y: List[float] = []
    origin_z: List[float] = []
    mags: List[float] = []
    if all_points:
        z_values = [float(p["z_mm"]) for p in all_points]
        z_min = min(min(z_values), 0.0)
        z_max = max(max(z_values), 0.0)
    else:
        z_min = -1.0
        z_max = 1.0
    z_pad = max(2.0, 0.06 * max(1.0, z_max - z_min))

    line_groups = [{"x": [], "y": [], "z": []} for _ in range(COLOR_BINS)]
    tip_groups = [{"x": [], "y": [], "z": []} for _ in range(COLOR_BINS)]

    for p in all_points:
        x0 = float(p["x_mm"])
        y0 = float(p["y_mm"])
        z0 = float(p["z_mm"])
        bx = float(p["bx_uT"])
        by = float(p["by_uT"])
        bz = float(p["bz_uT"])
        mag = math.sqrt(bx * bx + by * by + bz * bz)
        origin_x.append(x0)
        origin_y.append(y0)
        origin_z.append(z0)
        mags.append(mag)
        if mag <= 0.0:
            continue
        ux = bx / mag
        uy = by / mag
        uz = bz / mag
        x1 = x0 + ux * PLOT_ARROW_LENGTH_MM
        y1 = y0 + uy * PLOT_ARROW_LENGTH_MM
        z1 = z0 + uz * PLOT_ARROW_LENGTH_MM
        frac = min(mag, PLOT_MAG_MAX_UT) / PLOT_MAG_MAX_UT
        bin_idx = min(COLOR_BINS - 1, int(frac * COLOR_BINS))
        line_groups[bin_idx]["x"].extend([x0, x1, None])
        line_groups[bin_idx]["y"].extend([y0, y1, None])
        line_groups[bin_idx]["z"].extend([z0, z1, None])
        tip_groups[bin_idx]["x"].append(x1)
        tip_groups[bin_idx]["y"].append(y1)
        tip_groups[bin_idx]["z"].append(z1)

    data = []
    cx, cy, cz = reference_circle_xyz(PLOT_CIRCLE_RADIUS_MM)
    data.append({
        "type": "scatter3d",
        "mode": "lines",
        "showlegend": False,
        "x": cx,
        "y": cy,
        "z": cz,
        "line": {"color": "black", "width": 6},
        "hoverinfo": "skip",
    })
    data.append({
        "type": "scatter3d",
        "mode": "markers",
        "showlegend": False,
        "x": origin_x,
        "y": origin_y,
        "z": origin_z,
        "marker": {"size": PLOT_ORIGIN_MARKER_SIZE, "color": "rgba(230,230,230,0.95)"},
        "hovertemplate": "x=%{x:.2f} mm<br>y=%{y:.2f} mm<br>z=%{z:.2f} mm<extra></extra>",
    })
    for bin_idx in range(COLOR_BINS):
        frac = (bin_idx + 0.5) / COLOR_BINS
        color = color_for_fraction(frac)
        if line_groups[bin_idx]["x"]:
            data.append({
                "type": "scatter3d",
                "mode": "lines",
                "showlegend": False,
                "x": line_groups[bin_idx]["x"],
                "y": line_groups[bin_idx]["y"],
                "z": line_groups[bin_idx]["z"],
                "line": {"color": color, "width": PLOT_LINE_WIDTH_PX},
                "hoverinfo": "skip",
            })
        if tip_groups[bin_idx]["x"]:
            data.append({
                "type": "scatter3d",
                "mode": "markers",
                "showlegend": False,
                "x": tip_groups[bin_idx]["x"],
                "y": tip_groups[bin_idx]["y"],
                "z": tip_groups[bin_idx]["z"],
                "marker": {"size": PLOT_TIP_MARKER_SIZE, "color": color},
                "hoverinfo": "skip",
            })

    mag_min = min(mags) if mags else 0.0
    mag_max = max(mags) if mags else 0.0
    layout = {
        "title": {"text": title, "font": {"size": 18, "color": "#dddddd"}},
        "margin": {"l": 0, "r": 0, "b": 0, "t": 40},
        "paper_bgcolor": "#111111",
        "plot_bgcolor": "#111111",
        "scene": {
            "xaxis": {"title": "X (mm)", "range": [-PLOT_RADIUS_MM, PLOT_RADIUS_MM], "backgroundcolor": "#111111", "gridcolor": "#444444", "zerolinecolor": "#555555", "color": "#cccccc"},
            "yaxis": {"title": "Y (mm)", "range": [-PLOT_RADIUS_MM, PLOT_RADIUS_MM], "backgroundcolor": "#111111", "gridcolor": "#444444", "zerolinecolor": "#555555", "color": "#cccccc"},
            "zaxis": {"title": "Z (mm)", "range": [z_min - z_pad, z_max + z_pad], "backgroundcolor": "#111111", "gridcolor": "#444444", "zerolinecolor": "#555555", "color": "#cccccc"},
            "aspectmode": "manual",
            "aspectratio": {"x": 1, "y": 1, "z": max(0.35, (z_max - z_min + 2.0 * z_pad) / (2.0 * PLOT_RADIUS_MM))},
            "dragmode": "orbit",
        },
        "uirevision": "keep-camera",
        "annotations": [{
            "text": f"|B| range {mag_min:.1f}–{mag_max:.1f} uT, arrows fixed at 1 mm, color = magnitude",
            "xref": "paper",
            "yref": "paper",
            "x": 0.01,
            "y": 0.98,
            "showarrow": False,
            "font": {"size": 14, "color": "#dddddd"},
            "bgcolor": "rgba(0,0,0,0.15)",
        }],
    }
    return {"status": status, "data": data, "layout": layout}


class SilentLogHandler(SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return


class WebViewer:
    def __init__(self, port: int, auto_open: bool) -> None:
        self.port = port
        self.auto_open = auto_open
        self.dir = WEB_STATE_DIR
        self.dir.mkdir(parents=True, exist_ok=True)
        atomic_write_text(self.dir / "index.html", INDEX_HTML)
        write_json(self.dir / "state.json", {"status": "Viewer waiting for data…", "data": [], "layout": {}})
        handler = lambda *args, **kwargs: SilentLogHandler(*args, directory=str(self.dir), **kwargs)  # noqa: E731
        self.server = ThreadingHTTPServer(("127.0.0.1", port), handler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()
        self.url = f"http://127.0.0.1:{port}/index.html"
        print(f"Web viewer: {self.url}")
        if auto_open:
            try:
                webbrowser.open(self.url, new=1, autoraise=True)
            except Exception:
                pass

    def update(self, all_points: Sequence[Dict[str, object]], title: str, status: str) -> None:
        write_json(self.dir / "state.json", build_plotly_payload(all_points, title, status))

    def write_final_html(self, path: pathlib.Path, all_points: Sequence[Dict[str, object]], title: str, status: str) -> None:
        payload = build_plotly_payload(all_points, title, status)
        html = f"""<!doctype html>
<html><head><meta charset=\"utf-8\" /><meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
<title>MLX Final Field Map</title>
<script src=\"https://cdn.plot.ly/plotly-2.35.2.min.js\"></script>
<style>html,body,#plot{{margin:0;padding:0;width:100%;height:100%;background:#111;color:#ddd;font-family:Arial,sans-serif;}}</style>
</head><body><div id=\"plot\"></div>
<script>const payload = {json.dumps(payload)};
Plotly.newPlot('plot', payload.data || [], payload.layout || {{}}, {{responsive:true, scrollZoom:true, displaylogo:false}});</script>
</body></html>
"""
        atomic_write_text(path, html)
        if self.auto_open:
            try:
                webbrowser.open(path.resolve().as_uri(), new=1, autoraise=True)
            except Exception:
                pass

    def close(self) -> None:
        try:
            self.server.shutdown()
        except Exception:
            pass
        try:
            self.server.server_close()
        except Exception:
            pass


class PlotManager:
    def __init__(self, backend: str, auto_open: bool, port: int) -> None:
        self.backend = backend
        self.auto_open = auto_open
        self.web: Optional[WebViewer] = None
        self.all_points: List[Dict[str, object]] = []
        self.last_title = ""
        self.last_status = ""
        if backend == "web":
            self.web = WebViewer(port=port, auto_open=auto_open)

    def add_points(self, points: Sequence[Dict[str, object]]) -> None:
        self.all_points.extend(points)

    def update(self, title: str, status: str) -> None:
        self.last_title = title
        self.last_status = status
        if self.web is not None:
            self.web.update(self.all_points, title=title, status=status)

    def write_final_html(self, path: pathlib.Path) -> None:
        if self.web is not None:
            self.web.write_final_html(path, self.all_points, self.last_title or "Final field map", self.last_status)

    def close(self) -> None:
        if self.web is not None:
            self.web.close()


def print_capture_summary(prefix: str, wide_row: Dict[str, object]) -> None:
    print(
        f"{prefix} Captured snapshot @ {wide_row['snapshot_arduino_ms']} ms; "
        f"present={wide_row['sensors_present']}, measured={wide_row['sensors_measured']}"
    )


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Run an MLX90394 z-stack scan and live 3D field viewer.")
    p.add_argument("--mlx-source", default=None, help="MLX source: COM7 or tcp://127.0.0.1:8765")
    p.add_argument("--mlx-port", default=None, help="Backward-compatible alias for --mlx-source")
    p.add_argument("--stage-port", default=STAGE_PORT_DEFAULT, help=f"Stepper COM port (default: {STAGE_PORT_DEFAULT})")
    p.add_argument("--baud", type=int, default=DEFAULT_BAUD, help=f"Baud rate for serial sources (default: {DEFAULT_BAUD})")
    p.add_argument("--start-mm", type=float, default=Z_START_MM_DEFAULT, help=f"Start z in mm (default: {Z_START_MM_DEFAULT})")
    p.add_argument("--stop-mm", type=float, default=Z_STOP_MM_DEFAULT, help=f"Stop z in mm (default: {Z_STOP_MM_DEFAULT})")
    p.add_argument("--step-mm", type=float, default=Z_STEP_MM_DEFAULT, help=f"Step size in mm (default: {Z_STEP_MM_DEFAULT})")
    p.add_argument("--mapping", type=pathlib.Path, default=DEFAULT_MAPPING_XLSX, help="Path to sensor mapping workbook")
    p.add_argument("--state-file", type=pathlib.Path, default=DEFAULT_STATE_PATH, help="Stage state JSON path")
    p.add_argument("--output", type=pathlib.Path, default=DEFAULT_OUTPUT_CSV, help="Wide output CSV path")
    p.add_argument("--progress-json", type=pathlib.Path, default=DEFAULT_PROGRESS_JSON, help="Progress JSON path")
    p.add_argument("--raw-log-dir", type=pathlib.Path, default=None, help="Optional per-snapshot raw text log directory")
    p.add_argument("--single-measurement", action="store_true", help="Capture exactly one plane and plot it immediately")
    p.add_argument("--dummy-data", action="store_true", help="Use the dummy source and do not touch the motor")
    p.add_argument("--plot-backend", choices=("web", "none"), default="web", help="Plot backend (default: web)")
    p.add_argument("--web-port", type=int, default=WEB_PORT_DEFAULT, help=f"Localhost viewer port (default: {WEB_PORT_DEFAULT})")
    p.add_argument("--web-no-open", action="store_true", help="Do not auto-open the browser viewer")
    return p.parse_args()


def resolve_mlx_source(args: argparse.Namespace) -> str:
    src = args.mlx_source or args.mlx_port
    if src:
        return src
    return DUMMY_SOURCE_DEFAULT if args.dummy_data else MLX_SOURCE_DEFAULT


def run_single_measurement_real(args, mapping, output_csv, plotter, mlx_source_name, fieldnames) -> int:
    state = load_stage_state(args.state_file)
    z_steps = int(state.current_steps)
    z_um = target_um_from_steps_local(z_steps)
    z_mm = z_um / 1000.0
    print(f"Single-measurement mode at current z = {z_mm:.3f} mm ({z_steps} steps)")
    mlx = None
    try:
        mlx = open_source(mlx_source_name, args.baud, MLX_SERIAL_SETTLE_S)
        stop_mlx_stream(mlx)
        snapshot = capture_settled_snapshot(mlx, timeout_s=MLX_CAPTURE_TIMEOUT_S, raw_log_path=None)
        capture_ts = dt.datetime.now().astimezone()
        wide_row = build_wide_row(mapping, snapshot, capture_ts, z_index=0, z_steps=z_steps, z_um=z_um, z_mm=z_mm)
        append_wide_csv(output_csv, fieldnames, wide_row)
        print_capture_summary("[single]", wide_row)
        points = build_vector_points(mapping, snapshot.rows, z_mm=z_mm)
        plotter.add_points(points)
        plotter.update(title=f"Single measurement at z={z_mm:.3f} mm", status=f"{len(points)} vectors total, current z={z_mm:.3f} mm")
        print(f"Appended CSV row: {output_csv}")
        return 0
    finally:
        if mlx is not None:
            try:
                stop_mlx_stream(mlx)
            except Exception:
                pass
            try:
                mlx.close()
            except Exception:
                pass


def run_single_measurement_dummy(args, mapping, output_csv, plotter, mlx_source_name, fieldnames) -> int:
    z_steps = steps_from_mm_exact(args.start_mm, "start_mm")
    z_um = target_um_from_steps_local(z_steps)
    z_mm = z_um / 1000.0
    print(f"Single-measurement dummy mode at logical z = {z_mm:.3f} mm ({z_steps} steps)")
    mlx = None
    try:
        mlx = open_source(mlx_source_name, args.baud, MLX_SERIAL_SETTLE_S)
        set_dummy_z(mlx, z_mm)
        stop_mlx_stream(mlx)
        snapshot = capture_settled_snapshot(mlx, timeout_s=MLX_CAPTURE_TIMEOUT_S, raw_log_path=None)
        capture_ts = dt.datetime.now().astimezone()
        wide_row = build_wide_row(mapping, snapshot, capture_ts, z_index=0, z_steps=z_steps, z_um=z_um, z_mm=z_mm)
        append_wide_csv(output_csv, fieldnames, wide_row)
        print_capture_summary("[single dummy]", wide_row)
        points = build_vector_points(mapping, snapshot.rows, z_mm=z_mm)
        plotter.add_points(points)
        plotter.update(title=f"Dummy single measurement at z={z_mm:.3f} mm", status=f"{len(points)} vectors total, current z={z_mm:.3f} mm")
        print(f"Appended CSV row: {output_csv}")
        return 0
    finally:
        if mlx is not None:
            try:
                stop_mlx_stream(mlx)
            except Exception:
                pass
            try:
                mlx.close()
            except Exception:
                pass


def main() -> int:
    args = parse_args()
    mlx_source_name = resolve_mlx_source(args)
    mapping_path = args.mapping.resolve()
    state_path = args.state_file.resolve()
    output_csv = args.output.resolve()
    progress_json = args.progress_json.resolve()
    raw_log_dir = args.raw_log_dir.resolve() if args.raw_log_dir else None

    if not mapping_path.exists():
        print(f"Error: mapping workbook not found: {mapping_path}", file=sys.stderr)
        return 2
    if raw_log_dir is not None:
        raw_log_dir.mkdir(parents=True, exist_ok=True)

    try:
        mapping = load_position_map(mapping_path)
        fieldnames = csv_fieldnames()
        z_targets_steps = [] if args.single_measurement else build_z_targets_steps(args.start_mm, args.stop_mm, args.step_mm)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    plotter = PlotManager(backend=args.plot_backend, auto_open=not args.web_no_open, port=args.web_port)
    try:
        if args.single_measurement:
            if args.dummy_data:
                rc = run_single_measurement_dummy(args, mapping, output_csv, plotter, mlx_source_name, fieldnames)
            else:
                rc = run_single_measurement_real(args, mapping, output_csv, plotter, mlx_source_name, fieldnames)
            if rc == 0 and args.plot_backend == "web":
                final_html = output_csv.with_suffix(".html")
                plotter.write_final_html(final_html)
                print(f"Saved interactive plot: {final_html}")
            return rc

        print(f"Output CSV: {output_csv}")
        print(f"MLX source: {mlx_source_name}")
        if args.dummy_data:
            print("Mode: dummy data, no motor control")
        else:
            print(f"Stage port: {args.stage_port} @ {args.baud}")
        print(
            f"Z plan: {len(z_targets_steps)} positions from {args.start_mm:.3f} mm "
            f"to {args.stop_mm:.3f} mm in {args.step_mm:.3f} mm increments"
        )
        print(f"Discarding {MLX_SKIP_SNAPSHOTS_AFTER_START} snapshot(s) after each MLX restart")

        mlx = None
        stage_ser = None
        try:
            mlx = open_source(mlx_source_name, args.baud, MLX_SERIAL_SETTLE_S)
            if args.dummy_data:
                stop_mlx_stream(mlx)
                stage_api = None
            else:
                stage_api = load_stage_api()
                stage_ser = stage_api["open_stage_serial_port"](args.stage_port, args.baud, STAGE_SERIAL_SETTLE_S)
                _ = stage_api["drain_stage_serial"](stage_ser, STAGE_STARTUP_READ_S)
                stop_mlx_stream(mlx)

            total_positions = len(z_targets_steps)
            for z_index, target_steps in enumerate(z_targets_steps):
                target_um = target_um_from_steps_local(target_steps)
                target_mm = target_um / 1000.0
                if args.dummy_data:
                    print(f"[{z_index + 1}/{total_positions}] Dummy logical z = {target_mm:.3f} mm ({target_steps} steps)")
                    set_dummy_z(mlx, target_mm)
                else:
                    delta_steps, move_duration_s = move_stage_absolute(stage_ser, state_path, target_steps, timeout_s=STAGE_MOVE_TIMEOUT_S)
                    if delta_steps != 0:
                        print(
                            f"[{z_index + 1}/{total_positions}] Moved stage to {target_mm:.3f} mm "
                            f"({target_steps} steps, delta {delta_steps:+d}, {move_duration_s:.3f} s)"
                        )
                    else:
                        print(f"[{z_index + 1}/{total_positions}] Stage already at {target_mm:.3f} mm ({target_steps} steps)")
                    if STAGE_SETTLE_AFTER_MOVE_S > 0:
                        time.sleep(STAGE_SETTLE_AFTER_MOVE_S)

                raw_log_path = None
                if raw_log_dir is not None:
                    raw_log_path = raw_log_dir / f"snapshot_z{z_index:04d}_{target_steps:06d}steps.txt"

                snapshot = capture_settled_snapshot(mlx, timeout_s=MLX_CAPTURE_TIMEOUT_S, raw_log_path=raw_log_path)
                capture_ts = dt.datetime.now().astimezone()
                wide_row = build_wide_row(mapping, snapshot, capture_ts, z_index=z_index, z_steps=target_steps, z_um=target_um, z_mm=target_mm)
                append_wide_csv(output_csv, fieldnames, wide_row)
                print_capture_summary(f"[{z_index + 1}/{total_positions}]", wide_row)
                points = build_vector_points(mapping, snapshot.rows, z_mm=target_mm)
                plotter.add_points(points)
                status = (
                    f"{len(plotter.all_points)} vectors total, {len(points)} in most recent plane, "
                    f"|B| range 0–{PLOT_MAG_MAX_UT:.0f} uT color scale, current z={target_mm:.3f} mm"
                )
                title = f"Live field map: {z_index + 1}/{total_positions} planes captured, current z={target_mm:.3f} mm"
                plotter.update(title=title, status=status)
                write_json(
                    progress_json,
                    {
                        "completed_positions": z_index + 1,
                        "total_positions": total_positions,
                        "last_z_index": z_index,
                        "last_z_steps": target_steps,
                        "last_z_um": target_um,
                        "last_z_mm": target_mm,
                        "last_snapshot_arduino_ms": snapshot.snapshot_arduino_ms,
                        "last_capture_pc_local_time": capture_ts.isoformat(timespec="seconds"),
                        "output_csv": str(output_csv),
                        "mapping_xlsx": str(mapping_path),
                        "state_file": str(state_path),
                        "mlx_source": mlx_source_name,
                        "stage_port": None if args.dummy_data else args.stage_port,
                        "dummy_data": bool(args.dummy_data),
                    },
                )

            print(f"Completed z-stack. Output written to: {output_csv}")
            if args.plot_backend == "web":
                final_html = output_csv.with_suffix(".html")
                plotter.write_final_html(final_html)
                print(f"Saved interactive plot: {final_html}")
            return 0
        finally:
            if mlx is not None:
                try:
                    stop_mlx_stream(mlx)
                except Exception:
                    pass
                try:
                    mlx.close()
                except Exception:
                    pass
            if stage_ser is not None:
                try:
                    stage_ser.close()
                except Exception:
                    pass
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    finally:
        plotter.close()


if __name__ == "__main__":
    raise SystemExit(main())
